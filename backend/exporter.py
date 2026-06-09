"""Write collected rows to CSV / XLSX in the output dir and return a download ref.

The agent gathers rows during a run (the `record_rows` action) and writes a file
with the `export` action. xlsx needs openpyxl; if it's missing we fall back to csv
so a run never loses its data.
"""
from __future__ import annotations

import csv
import re

import config


def _safe_name(name: str, default: str = "export") -> str:
    base = re.sub(r"[^A-Za-z0-9_-]+", "_", (name or "").strip("/. ")).strip("_")
    return (base or default)[:80]


def _columns(rows: list[dict], columns) -> list[str]:
    if columns:
        return [str(c) for c in columns]
    seen: list[str] = []
    for r in rows:
        for k in r.keys():
            if str(k) not in seen:
                seen.append(str(k))
    return seen or ["value"]


def _cell(v):
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def _image_path(v):
    """If a cell value references a saved image in OUTPUT_DIR (a {"shot_of": N}
    result like '/output/img_5.png'), return its absolute Path so XLSX can embed
    the picture; else None (write it as plain text). Only files inside OUTPUT_DIR
    qualify — `Path(...).name` strips any path so a cell can't reach elsewhere."""
    if not isinstance(v, str):
        return None
    s = v.strip()
    if not re.search(r"\.(png|jpe?g)$", s, re.IGNORECASE):
        return None
    p = config.OUTPUT_DIR / s.rsplit("/", 1)[-1]
    return p if p.is_file() else None


def save_image(data: bytes, filename: str = "shot", ext: str = "png") -> dict:
    """Write screenshot bytes to output/<name>.<ext>, never overwriting an
    existing file. Returns {filename, url}."""
    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    name = _safe_name(filename, "shot")
    path = config.OUTPUT_DIR / f"{name}.{ext}"
    i = 2
    while path.exists():
        path = config.OUTPUT_DIR / f"{name}_{i}.{ext}"
        i += 1
    path.write_bytes(data)
    return {"filename": path.name, "url": f"/output/{path.name}"}


def write_table(rows: list[dict], filename: str, fmt: str = "xlsx", columns=None) -> dict:
    """Write `rows` to output/<name>.<fmt>. Returns {filename, rows, columns, url, format}."""
    rows = [r for r in (rows or []) if isinstance(r, dict)]
    fmt = (fmt or "xlsx").lower()
    if fmt not in ("xlsx", "csv"):
        fmt = "xlsx"
    cols = _columns(rows, columns)
    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    name = _safe_name(filename)

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
        except ImportError:
            fmt = "csv"  # graceful fallback so data is never lost

    path = config.OUTPUT_DIR / f"{name}.{fmt}"

    if fmt == "csv":
        # utf-8-sig so Excel opens non-ASCII correctly.
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            for r in rows:
                w.writerow({c: r.get(c, "") for c in cols})
    else:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        try:
            from openpyxl.drawing.image import Image as XLImage
        except Exception:  # noqa: BLE001 — no Pillow → degrade to path text
            XLImage = None

        wb = Workbook()
        ws = wb.active
        ws.append(cols)
        IMG_W = IMG_H = 130  # px thumbnail box embedded into the cell
        for ri, r in enumerate(rows, start=2):
            row_has_img = False
            for ci, c in enumerate(cols, start=1):
                v = r.get(c, "")
                img = _image_path(v)
                if img is not None and XLImage is not None:
                    try:
                        pic = XLImage(str(img))
                        pic.width, pic.height = IMG_W, IMG_H
                        ws.add_image(pic, f"{get_column_letter(ci)}{ri}")
                        col = ws.column_dimensions[get_column_letter(ci)]
                        col.width = max(col.width or 0, IMG_W / 7)  # px→Excel char-width
                        row_has_img = True
                        continue  # leave the cell empty; the picture floats over it
                    except Exception:  # noqa: BLE001 — embed failed → write the path
                        pass
                ws.cell(row=ri, column=ci, value=_cell(v))
            if row_has_img:
                ws.row_dimensions[ri].height = IMG_H * 0.75  # px→points
        wb.save(path)

    return {
        "filename": path.name,
        "rows": len(rows),
        "columns": cols,
        "url": f"/output/{path.name}",
        "format": fmt,
    }
